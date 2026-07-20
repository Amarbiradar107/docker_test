import os

from openpyxl import load_workbook

class ExcelUtil:
    @staticmethod
    def read_excel(filename,sheet_name):
        try:
            workbook = load_workbook(filename)
            sheet = workbook[sheet_name]
            testdata=[]
            row_count = sheet.max_row
            col_count = sheet.max_column
            for i in range(2,row_count):
                row =[]
                for j in range(1,col_count):
                    data = sheet.cell(row=i,column=j).value
                    row.append(data)
                testdata.append(row)
            return testdata
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None




        return sheet

